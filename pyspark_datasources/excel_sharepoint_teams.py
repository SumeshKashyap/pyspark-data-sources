import datetime
import logging
from collections.abc import Iterator
from dataclasses import dataclass
from io import BytesIO
from typing import Optional

from pyspark.sql.datasource import DataSource, DataSourceReader, InputPartition
from pyspark.sql.types import StringType, StructField, StructType

logger = logging.getLogger(__name__)

_GRAPH_API_BASE = "https://graph.microsoft.com/v1.0"
_GRAPH_SCOPE = "https://graph.microsoft.com/.default"
_REQUIRED_OPTIONS = [
    "tenant_id", "client_id", "client_secret", "site_host", "site_path", "file_path"
]


def _get_bearer_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    """Acquire a Microsoft Graph API bearer token using the client credentials flow."""
    try:
        from azure.identity import ClientSecretCredential
    except ImportError:
        raise ImportError(
            "azure-identity is required for the SharePoint Excel datasource. "
            "Install it with: pip install pyspark-data-sources[sharepoint-excel]"
        )
    try:
        credential = ClientSecretCredential(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret,
        )
        token = credential.get_token(_GRAPH_SCOPE)
        return token.token
    except Exception as exc:
        raise ConnectionError(f"Authentication failed: {exc}") from exc


def _resolve_site_id(token: str, site_host: str, site_path: str) -> str:
    """Resolve a SharePoint site ID from its host and path using Microsoft Graph API."""
    try:
        import requests
    except ImportError:
        raise ImportError("requests library is required.")

    url = f"{_GRAPH_API_BASE}/sites/{site_host}:{site_path}:"
    try:
        response = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    except requests.exceptions.ConnectionError as exc:
        raise ConnectionError(f"Network error contacting Microsoft Graph API: {exc}") from exc
    except requests.exceptions.Timeout:
        raise TimeoutError("Request to Microsoft Graph API timed out.")

    if response.status_code == 401:
        raise PermissionError(
            "Authentication rejected by Microsoft Graph: "
            "check tenant_id, client_id, and client_secret."
        )
    if response.status_code == 403:
        raise PermissionError(
            f"Access denied to SharePoint site '{site_host}{site_path}'. "
            "Ensure the app has Sites.Selected (with site grant) or Sites.Read.All permission."
        )
    if response.status_code == 404:
        raise FileNotFoundError(
            f"SharePoint site not found: '{site_host}{site_path}'. "
            "Verify that site_host and site_path are correct."
        )
    try:
        response.raise_for_status()
    except Exception as exc:
        raise ConnectionError(
            f"Microsoft Graph API error ({response.status_code}): {exc}"
        ) from exc

    return response.json()["id"]


def _download_file(token: str, site_id: str, file_path: str) -> bytes:
    """Download a file from SharePoint as raw bytes using Microsoft Graph API."""
    try:
        import requests
    except ImportError:
        raise ImportError("requests library is required.")

    url = f"{_GRAPH_API_BASE}/sites/{site_id}/drive/root:{file_path}:/content"
    try:
        response = requests.get(
            url, headers={"Authorization": f"Bearer {token}"}, timeout=120
        )
    except requests.exceptions.ConnectionError as exc:
        raise ConnectionError(f"Network error downloading file from SharePoint: {exc}") from exc
    except requests.exceptions.Timeout:
        raise TimeoutError(f"Download of '{file_path}' timed out.")

    if response.status_code == 403:
        raise PermissionError(
            f"Access denied to file '{file_path}'. "
            "Check that the app has Files.Read.All or that Sites.Selected grants file access."
        )
    if response.status_code == 404:
        raise FileNotFoundError(
            f"File not found in SharePoint: '{file_path}'. "
            "Verify that file_path is correct (e.g., '/General/data.xlsx')."
        )
    try:
        response.raise_for_status()
    except Exception as exc:
        raise ConnectionError(
            f"Failed to download '{file_path}' ({response.status_code}): {exc}"
        ) from exc

    return response.content


def _read_excel(
    content: bytes,
    has_header: bool,
    sheet_name: Optional[str] = None,
    header_row: Optional[str] = None,
    start_row: Optional[str] = None,
    start_column: Optional[str] = None,
    use_columns: Optional[str] = None,
    nrows: Optional[str] = None,
) -> tuple:
    """Parse Excel bytes with openpyxl and return ``(headers, data_rows)``.

    ``headers`` is a list of column name strings when ``has_header=True``,
    or ``None`` when ``has_header=False`` (caller relies on the Spark schema for names).
    ``data_rows`` is a list of tuples containing Python-native cell values.
    All option values arrive as strings from Spark options and are cast internally.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for the SharePoint Excel datasource. "
            "Install it with: pip install pyspark-data-sources[sharepoint-excel]"
        )

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                available = wb.sheetnames
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {available}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active
        all_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    # Skip leading rows (e.g., metadata rows before the header)
    skip = int(start_row) if start_row else 0
    if skip:
        all_rows = all_rows[skip:]

    # Extract or skip header
    header_idx = int(header_row) if header_row else 0
    if has_header:
        if header_idx >= len(all_rows):
            raise ValueError(
                f"header_row={header_idx} exceeds available rows ({len(all_rows)}) "
                "after applying start_row."
            )
        raw_header = all_rows[header_idx]
        headers: Optional[list] = [
            str(h) if h is not None else f"_col_{i}" for i, h in enumerate(raw_header)
        ]
        data_rows = list(all_rows[header_idx + 1:])
    else:
        headers = None
        data_rows = list(all_rows)

    # Apply start_column (drop first N columns)
    col_start = int(start_column) if start_column else 0
    if col_start:
        if headers is not None:
            headers = headers[col_start:]
        data_rows = [row[col_start:] for row in data_rows]

    # Apply use_columns (filter by header name; comma-separated list)
    if use_columns is not None and headers is not None:
        requested = [c.strip() for c in use_columns.split(",")]
        indices = [i for i, h in enumerate(headers) if h in requested]
        if not indices:
            raise ValueError(
                f"None of the requested columns {requested} "
                f"were found in sheet headers {headers}."
            )
        headers = [headers[i] for i in indices]
        data_rows = [tuple(row[i] if i < len(row) else None for i in indices) for row in data_rows]

    # Apply row limit
    if nrows is not None:
        data_rows = data_rows[: int(nrows)]

    return headers, data_rows


def _to_str(v: object) -> Optional[str]:
    """Convert a cell value to string, preserving None for empty cells."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return str(v)


def _validate_required_options(options: dict) -> None:
    """Raise ValueError listing any missing required options."""
    missing = [name for name in _REQUIRED_OPTIONS if not options.get(name)]
    if missing:
        raise ValueError(
            f"Missing required option(s): {', '.join(missing)}. "
            "Provide them via .option() in your read query."
        )


@dataclass
class SharePointExcelPartition(InputPartition):
    """Single-partition marker for a SharePoint Excel read operation."""

    pass


class SharePointExcelDataSource(DataSource):
    """
    A read-only DataSource for reading Excel files from SharePoint Online or Microsoft Teams.

    Connects to Microsoft Graph API using OAuth2 Client Credentials (Service Principal)
    to download the Excel file directly into memory (no temporary files on disk),
    then converts it to a Spark DataFrame via openpyxl.

    Microsoft Teams document libraries are backed by SharePoint sites, so reading a Teams
    channel file simply requires using the Teams site path (e.g., ``/teams/MyTeam``).

    Name: ``sharepoint_excel``

    Python Dependencies
    -------------------
    .. code-block:: bash

        pip install pyspark-data-sources[sharepoint-excel]

    Installs: ``azure-identity``, ``openpyxl``.

    Authentication Flow
    -------------------
    1. Acquire a bearer token from Azure AD using the Client Credentials flow.
    2. Resolve the SharePoint Site ID from ``site_host`` + ``site_path`` via Graph API.
    3. Download the Excel file as a byte stream from the site's default drive.
    4. Parse the byte stream with openpyxl and return a Spark DataFrame.

    Microsoft Graph Permissions
    ---------------------------
    The registered Azure AD application requires one of:

    - ``Sites.Selected`` (preferred — least-privilege) with an explicit site grant, **or**
    - ``Sites.Read.All``

    Sites.Selected Setup
    --------------------
    After assigning ``Sites.Selected`` in Azure AD (without global admin consent), a
    SharePoint administrator must grant the app read access to the target site:

    .. code-block:: bash

        POST https://graph.microsoft.com/v1.0/sites/{site-id}/permissions
        Content-Type: application/json

        {
            "roles": ["read"],
            "grantedToIdentities": [
                {
                    "application": {
                        "id": "<client-id>",
                        "displayName": "<app-display-name>"
                    }
                }
            ]
        }

    Options
    -------
    tenant_id : str
        Azure AD tenant ID.
    client_id : str
        Azure AD application (client) ID.
    client_secret : str
        Azure AD application client secret.
    site_host : str
        SharePoint host (e.g., ``contoso.sharepoint.com``).
    site_path : str
        SharePoint site path (e.g., ``/sites/MySharePoint`` or ``/teams/MyTeam``).
    file_path : str
        Path to the Excel file relative to the default drive root
        (e.g., ``/General/data.xlsx``).
    has_header : str, optional
        Whether the first data row is a header row. Default ``true``.
        Set to ``false`` when the sheet has no header; a custom schema must then be
        provided via ``.schema(...)``.
    sheet_name : str, optional
        Name of the sheet to read. Defaults to the workbook's *active* sheet
        (``openpyxl``'s ``Workbook.active``), which reflects whichever sheet was
        selected the last time the file was saved in Excel — not necessarily the
        first sheet by tab order. Pass ``sheet_name`` explicitly for deterministic reads.
    header_row : str, optional
        0-indexed row (within the sheet, after ``start_row`` is applied) to use as
        column headers. Default ``0``.
    start_row : str, optional
        Number of rows to skip at the top of the sheet before locating the header.
        Useful for skipping metadata rows. Default ``0``.
    start_column : str, optional
        0-indexed column position to start reading from; earlier columns are discarded.
        Default ``0``.
    use_columns : str, optional
        Comma-separated list of column names to read (e.g., ``"Name,Age,Revenue"``).
        Column names are matched against the header row.
    nrows : str, optional
        Maximum number of data rows to read.

    Schema
    ------
    By default, all columns are returned as ``StringType``.
    Use ``.schema(...)`` to override with typed columns, or when ``has_header=false``.

    Examples
    --------
    Register the data source:

    >>> from pyspark_datasources import SharePointExcelDataSource
    >>> spark.dataSource.register(SharePointExcelDataSource)

    Read an Excel file from SharePoint Online (all columns as strings):

    >>> df = (
    ...     spark.read.format("sharepoint_excel")
    ...     .option("tenant_id", "<tenant-id>")
    ...     .option("client_id", "<client-id>")
    ...     .option("client_secret", "<client-secret>")
    ...     .option("site_host", "contoso.sharepoint.com")
    ...     .option("site_path", "/sites/MySharePoint")
    ...     .option("file_path", "/General/data.xlsx")
    ...     .load()
    ... )
    >>> df.show()

    Read a sheet without a header row (custom schema required):

    >>> from pyspark.sql.types import StructType, StructField, StringType, LongType
    >>> schema = StructType([
    ...     StructField("name", StringType()),
    ...     StructField("age", LongType()),
    ... ])
    >>> df = (
    ...     spark.read.format("sharepoint_excel")
    ...     .schema(schema)
    ...     .option("tenant_id", "<tenant-id>")
    ...     .option("client_id", "<client-id>")
    ...     .option("client_secret", "<client-secret>")
    ...     .option("site_host", "contoso.sharepoint.com")
    ...     .option("site_path", "/sites/MySharePoint")
    ...     .option("file_path", "/General/data.xlsx")
    ...     .option("has_header", "false")
    ...     .load()
    ... )
    >>> df.show()

    Read a specific sheet with a row limit from a Teams channel:

    >>> df = (
    ...     spark.read.format("sharepoint_excel")
    ...     .option("tenant_id", "<tenant-id>")
    ...     .option("client_id", "<client-id>")
    ...     .option("client_secret", "<client-secret>")
    ...     .option("site_host", "contoso.sharepoint.com")
    ...     .option("site_path", "/teams/DataTeam")
    ...     .option("file_path", "/General/report.xlsx")
    ...     .option("sheet_name", "Q1 Report")
    ...     .option("nrows", "500")
    ...     .load()
    ... )
    >>> df.show()

    Read a sheet with the header on a later row and a custom typed schema restricted
    to a column range (e.g. header on the 3rd row, data in columns C:G):

    >>> from pyspark.sql.types import (
    ...     StructType, StructField, IntegerType, StringType, LongType, DoubleType
    ... )
    >>> schema = StructType([
    ...     StructField("ID", IntegerType()),
    ...     StructField("Name", StringType()),
    ...     StructField("Department", StringType()),
    ...     StructField("Salary", LongType()),
    ...     StructField("hike", DoubleType()),
    ... ])
    >>> df = (
    ...     spark.read.format("sharepoint_excel")
    ...     .schema(schema)
    ...     .option("tenant_id", "<tenant-id>")
    ...     .option("client_id", "<client-id>")
    ...     .option("client_secret", "<client-secret>")
    ...     .option("site_host", "contoso.sharepoint.com")
    ...     .option("site_path", "/sites/MySharePoint")
    ...     .option("file_path", "/General/data.xlsx")
    ...     .option("sheet_name", "Sheet2")
    ...     .option("start_row", "2")       # header sits on the 3rd row; skip the 2 rows above it
    ...     .option("header_row", "0")
    ...     .option("start_column", "2")    # column C is 0-indexed position 2
    ...     .option("use_columns", "ID,Name,Department,Salary,hike")  # caps the range at column G
    ...     .load()
    ... )
    >>> df.show()

    Write the result to Parquet or Delta Lake:

    >>> df.write.parquet("/path/to/output.parquet")
    >>> df.write.format("delta").save("/path/to/delta-table")
    """

    @classmethod
    def name(cls) -> str:
        return "sharepoint_excel"

    @property
    def _has_header(self) -> bool:
        return self.options.get("has_header", "true").lower() != "false"

    def schema(self) -> StructType:
        """Infer schema from the Excel header row. All columns are returned as StringType."""
        if not self._has_header:
            raise ValueError(
                "A custom schema must be provided when has_header='false'. "
                "Use .schema(...) to specify column names and types."
            )
        _validate_required_options(self.options)
        content = self._fetch_excel_bytes()
        headers, _ = _read_excel(
            content=content,
            has_header=True,
            sheet_name=self.options.get("sheet_name"),
            header_row=self.options.get("header_row"),
            start_row=self.options.get("start_row"),
            start_column=self.options.get("start_column"),
            use_columns=self.options.get("use_columns"),
            nrows="0",
        )
        return StructType([StructField(col, StringType()) for col in headers])

    def reader(self, schema: StructType) -> "SharePointExcelReader":
        return SharePointExcelReader(self.options, schema)

    def _fetch_excel_bytes(self) -> bytes:
        """Authenticate with Graph API and download the Excel file as bytes."""
        logger.info(
            "Acquiring Microsoft Graph API token for tenant '%s'",
            self.options["tenant_id"],
        )
        token = _get_bearer_token(
            self.options["tenant_id"],
            self.options["client_id"],
            self.options["client_secret"],
        )

        logger.info(
            "Resolving SharePoint site ID for %s%s",
            self.options["site_host"],
            self.options["site_path"],
        )
        site_id = _resolve_site_id(token, self.options["site_host"], self.options["site_path"])

        logger.info("Downloading Excel file: %s", self.options["file_path"])
        return _download_file(token, site_id, self.options["file_path"])


class SharePointExcelReader(DataSourceReader):
    """
    DataSourceReader for :class:`SharePointExcelDataSource`.

    Authenticates with Microsoft Graph API, downloads the Excel file into a
    ``BytesIO`` buffer (no disk writes), parses it with openpyxl, and yields rows
    matching the schema. All cell values are converted to strings for the default
    all-string schema; native Python types are passed through for custom typed schemas.
    """

    def __init__(self, options: dict, schema: StructType) -> None:
        self.options = options
        self.schema = schema

    def partitions(self) -> list:
        return [SharePointExcelPartition()]

    def read(self, partition: SharePointExcelPartition) -> Iterator[tuple]:
        has_header = self.options.get("has_header", "true").lower() != "false"

        logger.info("Acquiring Microsoft Graph API token")
        token = _get_bearer_token(
            self.options["tenant_id"],
            self.options["client_id"],
            self.options["client_secret"],
        )

        logger.info(
            "Resolving SharePoint site ID for %s%s",
            self.options["site_host"],
            self.options["site_path"],
        )
        site_id = _resolve_site_id(token, self.options["site_host"], self.options["site_path"])

        logger.info("Downloading Excel file: %s", self.options["file_path"])
        content = _download_file(token, site_id, self.options["file_path"])

        _, data_rows = _read_excel(
            content=content,
            has_header=has_header,
            sheet_name=self.options.get("sheet_name"),
            header_row=self.options.get("header_row"),
            start_row=self.options.get("start_row"),
            start_column=self.options.get("start_column"),
            use_columns=self.options.get("use_columns"),
            nrows=self.options.get("nrows"),
        )

        # Convert all values to strings for the default all-string schema.
        # For user-provided typed schemas, openpyxl returns Python-native types
        # (int, float, bool, datetime) which Spark maps directly.
        all_string = all(isinstance(f.dataType, StringType) for f in self.schema.fields)

        for row in data_rows:
            if all_string:
                yield tuple(_to_str(v) for v in row)
            else:
                yield tuple(row)
